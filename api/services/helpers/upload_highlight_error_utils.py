import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage, FileSystemStorage
from services.loggers.file_logger import FileLogger
from services.helpers.utils import Utils


class Service:
    @staticmethod
    def cleanup_file(file_path):
        if os.path.exists(file_path):
            os.remove(file_path)

    @staticmethod
    def save_wb(wb, file_path):
        wb.save(file_path)

    @staticmethod
    def get_s3_file_path(file_path):
        try:
            file_path_arr = file_path.split(".")
            ext = file_path_arr[-1]
            file_path_arr[-1] = str(Utils.now_timestamp()) + "." + ext
            return ".".join(file_path_arr)
        except Exception:
            return file_path

    @staticmethod
    def push_to_s3(wb, file_path):
        Service.save_wb(wb, file_path)

        fs = FileSystemStorage()
        file = fs.open(file_path)

        s3_file_path = Service.get_s3_file_path(file_path)

        default_storage.save(s3_file_path, ContentFile(file.read()))
        Service.cleanup_file(file_path)
        return default_storage.url(s3_file_path)

    @staticmethod
    def close_wb(wb, file_path):
        wb.close()
        Service.cleanup_file(file_path)


class UploadHighlightErrorUtils:
    @staticmethod
    def write_file(content, filename):
        file_path = os.path.join(settings.MEDIA_ROOT, "error_" + filename)
        Service.cleanup_file(file_path)

        fs = FileSystemStorage()
        fs.save(file_path, ContentFile(content))
        return file_path

    @staticmethod
    def load_ws(file_path):
        workbook = load_workbook(file_path)
        return workbook, workbook.worksheets[0]

    @staticmethod
    def highlight(ws, col_map):
        def inner(row_number, e):
            try:
                keys = list(e.get_codes().keys())
                col_indexes = [
                    col_map.get(key) for key in keys if col_map.get(key) is not None
                ]
                for col_index in col_indexes:
                    ws.cell(row=row_number, column=col_index + 1).fill = PatternFill(
                        start_color="FF0000", fill_type="solid"
                    )
            except Exception as e1:
                print(repr(e1))
                FileLogger.log(repr(e1), "upload")

        return inner

    def finish(wb, preview_file_path, num_errors):
        if num_errors:
            return Service.push_to_s3(wb, preview_file_path)
        Service.close_wb(wb, preview_file_path)
        return ""
